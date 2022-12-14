# Packages
import streamlit as st
import pandas as pd
import numpy as np
import warnings
import openpyxl as op
from io import BytesIO
from pyxlsb import open_workbook as open_xlsb
#from st_aggrid import AgGrid
#from st_aggrid import GridOptionsBuilder, AgGrid, GridUpdateMode, DataReturnMode

warnings.filterwarnings("ignore")

# App
def main():
    st.set_page_config(layout="wide")

    st.title('Excel Validator')

    st.info('**Information**')
    st.write('* **This app validates the number of columns, sum of numerical columns, count of rows and distinct categorical values**')
    st.write('* You have to import excel files with "xlsx" format')
    st.write("* Select the button  'Select Excel Files' and chose the file you want")
    st.write('* Select the sheet you want to compare')
    st.write("* Change the starting row if your sheet doesn't start in the first row")
    st.write("* Select the button 'Compare Excels' to validate the files")
    st.write('##')

    # Get excel files
    if st.checkbox('Select Excel Files', key=1):
        col1, col2 = st.columns(2)

        with col1:
            st.subheader('HAMS File')
            file_1 = st.file_uploader("Select HAMS file", type=['xlsx'])

            if file_1:
                comment_details='''file_details = {
                    "Filename": file_1.name,
                    "FileType": file_1.type,
                    "FileSize": file_1.size}'''

                st.write("##")
                wb = op.load_workbook(file_1)

                # Show Excel file
                #st.subheader("File details:")
                #st.json(file_details, expanded=False)
                st.markdown("----")

                # Skip Rows
                skip_rows_df_1 = st.number_input('Header Starting Row:', 0,100,1, key = 3)

                # Select sheet
                sheet_selector = st.selectbox("Select sheet:", wb.sheetnames, key='a')
                df_1 = pd.read_excel(file_1, sheet_selector, skiprows = skip_rows_df_1 - 1)
                st.write("##")
                st.markdown(f"### Currently Selected: `{sheet_selector}`")

                comment='''gb = GridOptionsBuilder.from_dataframe(df_1)
                gb.configure_pagination(paginationAutoPageSize=True)  # Add pagination
                gb.configure_side_bar()  # Add a sidebar
                gridOptions = gb.build()
                AgGrid(df_1,
                       gridOptions=gridOptions,
                       data_return_mode='AS_INPUT',
                       update_mode='MODEL_CHANGED',
                       fit_columns_on_grid_load=False,
                       theme='streamlit',
                       enable_enterprise_modules=True,
                       height=500,
                       width='100%',
                       reload_data=True)'''

                st.dataframe(df_1)

                # SUM
                st.write("##")
                #st.subheader('SUM of columns')
                df_1_sum = df_1.sum(numeric_only=True)
                #st.dataframe(df_1_sum)

                # Count
                st.write("##")
                #st.subheader('Count of columns')
                df_1_count = df_1.count(numeric_only=True)
                #st.dataframe(df_1_count)

        with col2:
            st.subheader('DWH File')
            file_2 = st.file_uploader("Select DWH file", type=['xlsx'])

            if file_2:
                comment_details='''file_details = {
                    "Filename": file_2.name,
                    "FileType": file_2.type,
                    "FileSize": file_2.size}'''

                st.write("##")
                wb = op.load_workbook(file_2)

                # Show Excel file
                #st.subheader("File details:")
                #st.json(file_details, expanded=False)
                st.markdown("----")

                # Skip Rows
                skip_rows_df_2 = st.number_input('Header Starting Row:', 0, 100, 1, key = 4)

                # Select sheet
                sheet_selector = st.selectbox("Select sheet:", wb.sheetnames, key='b')
                df_2 = pd.read_excel(file_2, sheet_selector, skiprows = skip_rows_df_2 - 1)
                st.write("##")
                st.markdown(f"### Currently Selected: `{sheet_selector}`")

                comment='''gb2 = GridOptionsBuilder.from_dataframe(df_2)
                gb2.configure_pagination(paginationAutoPageSize=True)  # Add pagination
                gb2.configure_side_bar()  # Add a sidebar
                gridOptions = gb2.build()
                AgGrid(df_2,
                       gridOptions=gridOptions,
                       data_return_mode='AS_INPUT',
                       update_mode='MODEL_CHANGED',
                       fit_columns_on_grid_load=False,
                       theme='streamlit',
                       enable_enterprise_modules=True,
                       height=500,
                       width='100%',
                       reload_data=True)'''

                st.dataframe(df_2)

                # SUM
                st.write("##")
                #st.subheader('SUM of columns')
                df_2_sum = df_2.sum(numeric_only=True)
                #st.dataframe(df_2_sum)

                # Count
                st.write("##")
                #st.subheader('Count of columns')
                df_2_count = df_2.count(numeric_only=True)
                #st.dataframe(df_2_count)

            st.write("##")

    ## Comparison
    # SUM
    st.write('##')
    if st.checkbox('Compare Excels', key=2):
        sum = 'SUM'
        cols = 'COLUMNS'

        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.subheader(f"Validation: `{cols}`")
            cols_df_1 = len(df_1.axes[1])
            cols_df_2 = len(df_2.axes[1])

            st.write('**HAMS**')
            st.text(f'{cols_df_1} columns')

            st.write('**DWH**')
            st.text(f'{cols_df_2} columns')

            st.write('##')

        with col2:
            st.subheader(f"Validation: `{sum}`")

            #df_unique = df_1.nunique().to_frame().reset_index()
            #st.dataframe(df_unique)

            # Series to Dataframe
            try:
                df_1_sum = df_1_sum.to_frame()
                df_2_sum = df_2_sum.to_frame()

                # Compare
                #df_compare_sum =df_1_sum.compare(df_2_sum, align_axis=0)
                st.write('**HAMS**')
                st.dataframe(df_1_sum)

                st.write('**DWH**')
                st.dataframe(df_2_sum)

                df_compare_sum = df_1_sum - df_2_sum

                st.write('**Diference**')
                st.dataframe(df_compare_sum)
                #st.dataframe(df_compare_sum)

                # Dif between sheets
                #st.markdown('#### Diference between sheets')
                #df_dif_sum = df_compare_sum.diff()
                df_dif_sum = df_compare_sum.sum()
                #st.dataframe(df_dif_sum)

                st.write("##")
            except:
                st.error('Sheets with different formats')

                cols_df_1 = len(df_1.axes[1])
                cols_df_2 = len(df_2.axes[1])

                # Number of Columns
                st.write(f'First file has **{cols_df_1}** columns and second file has **{cols_df_2}** columns')

        with col3:
            # Count
            count = 'COUNT'
            st.subheader(f"Validation: `{count}`")

            # Series to Dataframe
            try:
                df_1_count = df_1_count.to_frame()
                df_2_count = df_2_count.to_frame()

                # Compare
                #df_compare_count = df_1_count.compare(df_2_count, align_axis=0)

                st.write('**HAMS**')
                st.dataframe(df_1_count)

                st.write('**DWH**')
                st.dataframe(df_2_count)

                df_compare_count = df_1_count - df_2_count

                st.write('**Diference**')
                st.dataframe(df_compare_count)

                # Dif between sheets
                st.write('##')
                st.write('##')

                #st.markdown('#### **Diference between sheets**')
                #df_dif_count = df_compare_count.diff()
                df_dif_count = df_compare_count.sum()
                #st.dataframe(df_dif_count)

            except:
                st.error('Sheets with different formats')

                cols_df_1 = len(df_1.axes[1])
                cols_df_2 = len(df_2.axes[1])

                # Number of Columns
                st.write(f'First file has **{cols_df_1}** columns and second file has **{cols_df_2}** columns')


        with col4:
            # Count
            distinct = 'DISTINCT'
            st.subheader(f"Validation: `{distinct}`")

            st.write('**HAMS**')
            df_1_distinct = df_1.nunique(dropna=False).sum()
            st.write(f'{df_1_distinct} distinct categorical values')
            st.write(df_1.nunique().to_frame().reset_index())

            st.write('**DWH**')
            df_2_distinct = df_2.nunique(dropna=False).sum()
            st.write(f'{df_2_distinct} distinct categorical values')
            st.write(df_2.nunique().to_frame().reset_index())

            st.write('**Diference**')
            st.dataframe(df_1.nunique(dropna=False) - df_2.nunique(dropna=False))

        ## Final Validation

        # Final Dataframe
        #df_dif_sum = df_dif_sum.rename(index={2: "SUM"})
        #st.dataframe(df_dif_sum)
        #df_dif_count = df_dif_count.rename(index={0: "COUNT"})

        #df_validation = pd.append([df_dif_sum, df_dif_count], axis=1)
        df_validation = df_dif_sum.append(df_dif_count)
        #df_validation = df_dif_sum - df_dif_count
        #df_validation.rename(index={0: "SUM", 1: "COUNT"})

        # Diference between sheets
        new_title = '<p style="font-family:Arial; color:#008080; font-size: 35px;">Diference between Sheets</p>'
        st.markdown(new_title, unsafe_allow_html=True)
        #st.dataframe(df_validation, width = 100)

        # Drop NaN values (extra columns)
        #df_compare_sum = df_compare_sum.dropna()
        #df_compare_count = df_compare_count.dropna()

        validation_columns = cols_df_1 == cols_df_2
        validation_distinct = df_1_distinct == df_2_distinct

        # Check SUM
        if df_compare_sum.any().sum() > 0 :
            st.write('**SUM NOK**')
        else:
            st.write('**SUM OK**')

        # Check COUNT
        if df_compare_count.any().sum() > 0:
            st.write('**COUNT NOK**')
        else:
            st.write('**COUNT OK**')

        # Check COLUMNS
        if validation_columns == True:
            st.write('**COLUMNS OK**')
        else:
            st.write('**COLUMNS NOK**')

        # Check DISTINCT CATEGORICAL VALUES
        if df_1_distinct == df_2_distinct:
            st.write('**DISTINCT CATEGORICAL OK**')
        else:
            st.write('**DISTINCT CATEGORICAL NOK**')

        # Final Check
        st.write('##')
        if df_compare_sum.any().sum() > 0 or df_compare_count.any().sum() > 0 or validation_columns == False or validation_distinct == False:
            st.error('**Sheet Not Validated!**')
        else:
            st.success('**Sheet Validated!**')

        st.write("##")
        st.subheader('Download validation results')
        #st.download_button(label='???? Download Excel', data=df_validation.to_csv(), mime='text/csv')

        comment_download_excel='''def to_excel(df):
            output = BytesIO()
            writer = pd.ExcelWriter(output, engine='openpyxl')
            df.to_excel(writer, index=False, sheet_name='Validation')
            workbook = writer.book
            worksheet = writer.sheets['Validation']
            writer.save()
            processed_data = output.getvalue()
            return processed_data

        df_xlsx = to_excel(df_validation)
        st.download_button(label='???? Download Excel',
                           data=df_xlsx,
                           file_name='df_test.xlsx')'''


        df_download = df_compare_sum.append(df_compare_count)
        st.download_button(label='???? Download Excel', data=df_download.to_csv(), mime='text/csv')

if __name__ == '__main__':
    main()